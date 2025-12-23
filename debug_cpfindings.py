#!/usr/bin/env python3
"""
Debug script to examine workItem.xlsx structure and extract CPFindings data
"""

import openpyxl
import json
import os

def examine_excel_structure():
    """Examine the structure of workItem.xlsx"""
    excel_path = 'adminlte/aws/956288449190/workItem.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Excel file not found: {excel_path}")
        return
    
    print(f"=== EXAMINING EXCEL STRUCTURE ===")
    print(f"File: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    print(f"Sheets: {wb.sheetnames}")
    
    sheets_to_skip = ['Info', 'Appendix']
    
    for sheet_name in wb.sheetnames:
        if sheet_name in sheets_to_skip:
            print(f"\n--- SKIPPING {sheet_name} ---")
            continue
            
        ws = wb[sheet_name]
        print(f"\n--- SHEET: {sheet_name} ---")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        if ws.max_row > 0:
            # Print headers
            headers = []
            for col in range(1, min(ws.max_column + 1, 10)):  # Limit to first 10 columns
                header = ws.cell(1, col).value
                headers.append(header)
            print(f"Headers (first 10): {headers}")
            
            # Print first few data rows
            for row in range(2, min(ws.max_row + 1, 5)):  # First 3 data rows
                row_data = []
                for col in range(1, min(ws.max_column + 1, 10)):
                    cell_value = ws.cell(row, col).value
                    row_data.append(str(cell_value)[:50] if cell_value else '')
                print(f"Row {row}: {row_data}")

def extract_cpfindings_data():
    """Extract CPFindings data like OutputGenerator does"""
    excel_path = 'adminlte/aws/956288449190/workItem.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Excel file not found: {excel_path}")
        return
    
    print(f"\n=== EXTRACTING CPFINDINGS DATA ===")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    findings = []
    suppressed = []
    columns = []
    
    sheets_to_skip = ['Info', 'Appendix']
    
    # Get column headers from first sheet
    for sheet_name in wb.sheetnames:
        if sheet_name not in sheets_to_skip:
            ws = wb[sheet_name]
            if ws.max_row > 0:
                columns = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            break
    
    print(f"Columns: {columns}")
    
    # Extract data from all sheets
    for sheet_name in wb.sheetnames:
        if sheet_name in sheets_to_skip:
            continue
        
        ws = wb[sheet_name]
        
        # Skip if sheet is empty
        if ws.max_row < 2:
            continue
        
        print(f"\nProcessing sheet: {sheet_name} ({ws.max_row - 1} rows)")
        
        # Extract rows
        for row in range(2, ws.max_row + 1):
            finding = {'service': sheet_name}
            
            # Extract all columns
            for col_idx, header in enumerate(columns, 1):
                cell_value = ws.cell(row, col_idx).value
                # Convert None to empty string
                finding[header if header else f'Column{col_idx}'] = cell_value if cell_value is not None else ''
            
            # Separate by status (last column is typically Status)
            status = finding.get('Status', finding.get('status', ''))
            if status == 'Suppressed':
                suppressed.append(finding)
            else:
                findings.append(finding)
    
    print(f"\nExtracted {len(findings)} findings and {len(suppressed)} suppressed items")
    
    # Show sample findings
    if findings:
        print(f"\nSample finding:")
        sample = findings[0]
        for key, value in sample.items():
            print(f"  {key}: {value}")
    
    if suppressed:
        print(f"\nSample suppressed:")
        sample = suppressed[0]
        for key, value in sample.items():
            print(f"  {key}: {value}")
    
    return {
        'columns': columns,
        'findings': findings,
        'suppressed': suppressed
    }

if __name__ == "__main__":
    examine_excel_structure()
    data = extract_cpfindings_data()
    
    # Save to JSON for inspection
    with open('debug_cpfindings_output.json', 'w') as f:
        json.dump(data, f, indent=2, default=str)
    
    print(f"\nData saved to debug_cpfindings_output.json")